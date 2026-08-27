"""Download WHO growth reference tables and vendor them into the repo as CSV.

Run once (or when WHO republishes):   python scripts/fetch_who_tables.py

Why this exists
---------------
Section 6.4 of the master prompt requires that growth classification be
deterministic and auditable -- never model output. That means the reference
tables must live in the repository, with a traceable provenance record, rather
than being fetched at runtime or bundled inside a third-party package.

Each generated CSV carries a `#` header block naming the exact WHO source URL,
the retrieval date and the SHA-256 of the source workbook, so a reviewer can
re-derive any number in this system from a WHO artefact.

We keep WHO's own published SD3neg/SD2neg/SD2pos/SD3pos columns alongside L/M/S.
We do not use them at runtime -- we recompute them from LMS -- but the test
suite asserts our computed values reproduce WHO's published ones exactly. That
is what makes the "100% match against WHO Anthro" target in Section 6.5 a
checkable claim rather than an assertion.

Naming caveat, verified 2026-08-28: the two weight-for-age (5-10 years) files on
WHO's CDN are named `hfa-*` even though they contain weight-for-age data. The
sheet name inside (`wfa_boys_z_WHO 2007_exp`) and the M values (~18.5 kg at 61
months) confirm the contents. `verify_indicator` below asserts this at fetch
time so a future re-fetch cannot silently pick up the wrong file.
"""

from __future__ import annotations

import csv
import hashlib
import io
from datetime import date
from pathlib import Path

import httpx
import openpyxl

OUT_DIR = Path(__file__).resolve().parent.parent / "app" / "growth" / "who"
RAW_DIR = Path(__file__).resolve().parent / "_who_raw"

CGS = "https://cdn.who.int/media/docs/default-source/child-growth/child-growth-standards/indicators"
REF7 = "https://cdn.who.int/media/docs/default-source/child-growth/growth-reference-5-19-years"

# (output_stem, url, key_column_name, key_unit, plausible M range at first row)
SOURCES: list[tuple[str, str, str, str, tuple[float, float]]] = [
    # --- WHO Child Growth Standards 2006, 0-60 months ----------------------
    (
        "wfa_boys_0_60m",
        f"{CGS}/weight-for-age/expanded-tables/wfa-boys-zscore-expanded-tables.xlsx",
        "Day",
        "day",
        (2.5, 4.5),
    ),
    (
        "wfa_girls_0_60m",
        f"{CGS}/weight-for-age/expanded-tables/wfa-girls-zscore-expanded-tables.xlsx",
        "Day",
        "day",
        (2.5, 4.5),
    ),
    (
        "hfa_boys_0_60m",
        f"{CGS}/length-height-for-age/expandable-tables/lhfa-boys-zscore-expanded-tables.xlsx",
        "Day",
        "day",
        (45.0, 55.0),
    ),
    (
        "hfa_girls_0_60m",
        f"{CGS}/length-height-for-age/expandable-tables/lhfa-girls-zscore-expanded-tables.xlsx",
        "Day",
        "day",
        (45.0, 55.0),
    ),
    (
        "wfl_boys_45_110cm",
        f"{CGS}/weight-for-length-height/expanded-tables/wfl-boys-zscore-expanded-table.xlsx",
        "Length",
        "cm",
        (2.0, 3.0),
    ),
    (
        "wfl_girls_45_110cm",
        f"{CGS}/weight-for-length-height/expanded-tables/wfl-girls-zscore-expanded-table.xlsx",
        "Length",
        "cm",
        (2.0, 3.0),
    ),
    (
        "wfh_boys_65_120cm",
        f"{CGS}/weight-for-length-height/expanded-tables/wfh-boys-zscore-expanded-tables.xlsx",
        "Height",
        "cm",
        (6.5, 8.5),
    ),
    (
        "wfh_girls_65_120cm",
        f"{CGS}/weight-for-length-height/expanded-tables/wfh-girls-zscore-expanded-tables.xlsx",
        "Height",
        "cm",
        (6.5, 8.5),
    ),
    # --- WHO Growth Reference 2007, 5-19 years -----------------------------
    (
        "hfa_boys_61_228m",
        f"{REF7}/height-for-age-(5-19-years)/hfa-boys-z-who-2007-exp.xlsx",
        "Month",
        "month",
        (105.0, 115.0),
    ),
    (
        "hfa_girls_61_228m",
        f"{REF7}/height-for-age-(5-19-years)/hfa-girls-z-who-2007-exp.xlsx",
        "Month",
        "month",
        (105.0, 115.0),
    ),
    (
        "wfa_boys_61_120m",
        f"{REF7}/weight-for-age-(5-10-years)/hfa-boys-z-who-2007-exp_0ff9c43c-8cc0-4c23-9fc6-81290675e08b.xlsx",
        "Month",
        "month",
        (17.0, 20.0),
    ),
    (
        "wfa_girls_61_120m",
        f"{REF7}/weight-for-age-(5-10-years)/hfa-girls-z-who-2007-exp_7ea58763-36a2-436d-bef0-7fcfbadd2820.xlsx",
        "Month",
        "month",
        (17.0, 20.0),
    ),
    (
        "bfa_boys_61_228m",
        f"{REF7}/bmi-for-age-(5-19-years)/bmi-boys-z-who-2007-exp.xlsx",
        "Month",
        "month",
        (14.0, 17.0),
    ),
    (
        "bfa_girls_61_228m",
        f"{REF7}/bmi-for-age-(5-19-years)/bmi-girls-z-who-2007-exp.xlsx",
        "Month",
        "month",
        (14.0, 17.0),
    ),
]

WANTED_SD = ["SD3neg", "SD2neg", "SD2", "SD3"]
OUT_SD = ["sd3neg", "sd2neg", "sd2pos", "sd3pos"]


def verify_indicator(stem: str, first_m: float, expected: tuple[float, float]) -> None:
    """Guard against WHO's misnamed files silently swapping an indicator."""
    lo, hi = expected
    if not lo <= first_m <= hi:
        raise SystemExit(
            f"{stem}: first-row M={first_m} is outside the expected range "
            f"{lo}-{hi} for this indicator. The source file at this URL is "
            f"probably not the table we think it is -- inspect it before trusting it."
        )


def convert(stem: str, url: str, key_col: str, key_unit: str, expected: tuple[float, float]) -> int:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    raw = RAW_DIR / f"{stem}.xlsx"
    if raw.exists():
        blob = raw.read_bytes()
    else:
        resp = httpx.get(url, timeout=120, follow_redirects=True)
        resp.raise_for_status()
        blob = resp.content
        raw.write_bytes(blob)
    digest = hashlib.sha256(blob).hexdigest()

    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]

    if header[0] != key_col:
        raise SystemExit(f"{stem}: expected first column {key_col!r}, found {header[0]!r}")
    idx = {name: header.index(name) for name in ("L", "M", "S")}
    for name in WANTED_SD:
        if name not in header:
            raise SystemExit(f"{stem}: source is missing published column {name!r}")
        idx[name] = header.index(name)

    out_rows: list[list[object]] = []
    for row in rows:
        if row[0] is None or row[idx["M"]] is None:
            continue
        out_rows.append(
            [row[0], row[idx["L"]], row[idx["M"]], row[idx["S"]]] + [row[idx[n]] for n in WANTED_SD]
        )
    wb.close()
    if not out_rows:
        raise SystemExit(f"{stem}: no data rows parsed")

    verify_indicator(stem, float(out_rows[0][2]), expected)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = OUT_DIR / f"{stem}.csv"
    with out.open("w", newline="", encoding="utf-8") as fh:
        fh.write("# WHO growth reference table, vendored for PoshanNetra AI (Section 6.4).\n")
        fh.write(f"# source_url: {url}\n")
        fh.write(f"# retrieved: {date.today().isoformat()}\n")
        fh.write(f"# sha256(source_xlsx): {digest}\n")
        fh.write(f"# key column '{key_col}' is in {key_unit}s; rows: {len(out_rows)}\n")
        fh.write("# sd* columns are WHO's OWN published cut-offs, kept only so the test\n")
        fh.write("# suite can prove our LMS implementation reproduces them exactly.\n")
        w = csv.writer(fh)
        w.writerow(["key", "l", "m", "s", *OUT_SD])
        w.writerows(out_rows)
    return len(out_rows)


def main() -> None:
    total = 0
    for stem, url, key_col, key_unit, expected in SOURCES:
        n = convert(stem, url, key_col, key_unit, expected)
        total += n
        print(f"  {stem:24s} {n:5d} rows")
    print(f"\n{len(SOURCES)} tables, {total} rows -> {OUT_DIR}")


if __name__ == "__main__":
    main()
